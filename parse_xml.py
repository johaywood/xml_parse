import xml.etree.ElementTree as ET
from openpyxl import Workbook
import sys

class AutoVivification(dict):
    """Implementation of perl's autovivification feature"""
    def __getitem__(self, item):
        try:
            return dict.__getitem__(self, item)
        except KeyError:
            value = self[item] = type(self)()
            return value

#Generate tree in memory from XML file
tree = ET.parse(sys.argv[1])      ## Individual animal data file
root = tree.getroot()             ## Get root for individual animal data file
grp_tree = ET.parse(sys.argv[2])  ## Group data file
grp_root = grp_tree.getroot()     ## Get root for group data file

#Initialize lists and dicts
animal_uars = []                        ## Individual animal ID list
meas_ids = []                           ## Individual measurement ID list
meas_by_day = []                        ## Individual animal measurements by day (eg. Bodyweight Day 1, Bodyweight Day 6, etc)
grp_ids = []                            ## Individual animal file group IDs
grp_summary_ids = []                    ## Group statistics summary IDs (eg. Mean, SEM, N, etc)
group_grp_ids = []                      ## Group file group IDs
grp_meas_by_day = []                    ## Group animal measurements by day (eg. Bodyweight Day 1, Bodyweight Day 6, etc)
animal_key = AutoVivification()         ## {Animal UAR : Animal Name}
an_grp_key = AutoVivification()         ## {Animal UAR : Animal Group}
grp_key = AutoVivification()            ## Individual animal file {Group ID : Group Name}
group_grp_key = AutoVivification()      ## Group file {Group ID : Group Name}
meas_key = AutoVivification()           ## Individual animal file {Measurement ID : Measurement Name}
grp_summary_key = AutoVivification()    ## Group summary stats key {Summary ID : Summary Name} (eg. {17 : Mean})
results = AutoVivification()            ## Dictionary of individual animal results { UAR : { meas_by_day : { RESULT_STRING : result } } }
grp_results = AutoVivification()        ## Dictionary of group results { grp_meas_by_day : { group_grp_id : { grp_summary_id : { RESULT_STRING : result } } } }

study_number = str()
study_title = str()

#Collect study number and title
for sn in root.iter('STUDY'):
  study_number = str(sn.find('STUDY_REFERENCE').text)
  study_title = str(sn.find('STUDY_TITLE').text)

#Collect individual animal ID and group ID info
for anim in root.iter('ANIMAL'):
  anim_uar = str(anim.find('ANIMAL_UAR').text)
  anim_num = str(anim.find('ANIMAL_REFERENCE').text)
  anim_grp = str(anim.find('GROUP_ID').text)
  animal_uars.append(anim_uar)
  animal_key[anim_uar] = anim_num
  an_grp_key[anim_uar] = anim_grp

#Collect measurement ID info
for meas in root.iter('MEASUREMENT'):
  meas_id = str(meas.find('MEASUREMENT_ID').text)
  meas_name = str(meas.find('MEASUREMENT_DESCR').text)
  meas_ids.append(meas_id)
  meas_key[meas_id] = meas_name
  
#Collect individual file group ID info
for grp in root.iter('GROUP'):
  grp_id = str(grp.find('GROUP_ID').text)
  grp_name = str(grp.find('GROUP_LONG_NAME').text)
  grp_ids.append(grp_id)
  grp_key[grp_id] = grp_name
  
#Collect group file group ID info
for grp in grp_root.iter('GROUP'):
  grp_id = str(grp.find('GROUP_ID').text)
  grp_name = str(grp.find('GROUP_LONG_NAME').text)
  group_grp_ids.append(grp_id)
  group_grp_key[grp_id] = grp_name
  
#Collect group summary ID info
for summ in grp_root.iter('GROUP_SUMMARY'):
  summ_id = str(summ.find('GROUP_SUMMARY_ID').text)
  summ_name = str(summ.find('GROUP_SUMMARY_DESCR').text)
  grp_summary_ids.append(summ_id)
  grp_summary_key[summ_id] = summ_name
  
#Collect animal results by measurement and UAR
for result in root.iter('ANIMAL_RESULT'):
  for uar_result in result.iter('ANIMAL_UAR'):
    an = uar_result.text
  for mid_result in result.iter('MEASUREMENT_ID'):
    mid = mid_result.text
  for tp_from in result.iter('TIME_PERIOD_FROM'):
    tpf = tp_from.text
  for tp_to in result.iter('TIME_PERIOD_TO'):
    tpt = tp_to.text
  if tpf == tpt:
    tp = meas_key[mid] + ' Day ' + tpf ## Single day measurement
  else:
    tp = meas_key[mid] + ' Day ' + tpf + ' - Day ' + tpt ## Calculation over time measurement (eg. Day 1 - Day 6 change)
  meas_by_day.append(tp)
  for str_result in result.iter('RESULT_STRING'):
    val = str_result.text
    results[an][tp]['RESULT_STRING'] = val
meas_by_day = set(meas_by_day)
meas_by_day = list(meas_by_day)
  
#Collect group results by measurement, timepoint, group ID
for result in grp_root.iter('GROUP_SUMMARY_RESULT'):
  for grp_id_result in result.iter('GROUP_ID'):
    gn = grp_id_result.text
    group_grp_ids.append(gn)
  for mid_result in result.iter('MEASUREMENT_ID'):
    mid = mid_result.text
  for gsid_result in result.iter('GROUP_SUMMARY_ID'):
    gsid = gsid_result.text
  for tp_from in result.iter('TIME_PERIOD_FROM'):
    tpf = tp_from.text
  for tp_to in result.iter('TIME_PERIOD_TO'):
    tpt = tp_to.text
  if tpf == tpt:
    tp = meas_key[mid] + ' Day ' + tpf
  else:
    tp = meas_key[mid] + ' Day ' + tpf + ' - Day ' + tpt
  grp_meas_by_day.append(tp)  
  for grp_str_result in result.iter('GROUP_RESULT_STRING'):
    val = grp_str_result.text
    grp_results[tp][gn][gsid]['RESULT_STRING'] = val
grp_meas_by_day = set(grp_meas_by_day) 
grp_meas_by_day = list(grp_meas_by_day)
group_grp_ids = set(group_grp_ids)
group_grp_ids = list(group_grp_ids)
group_grp_ids.sort()

#Initialize openpyxl workbook in memory
wb = Workbook()
ws = wb.active
ws.title = study_number + " Ind. Animal Data"
ws2 = wb.create_sheet()
ws2.title = study_number + " Group Summary"

#Add animal # and measurement column headers from lookup of measurement descriptions
ws.cell(row=1, column=1).value = "Animal #"
ws.cell(row=1, column=2).value = "Group"
for j in range (1, len(meas_by_day) + 1):
  ws.cell(row=1, column=j+2).value = meas_by_day[j-1]

#Insert animal number and group name via lookup of UAR
for i in range(0, len(animal_uars)): 
  ws.cell(row=i+2, column=1).value = animal_key[animal_uars[i]]
  ws.cell(row=i+2, column=2).value = grp_key[an_grp_key[animal_uars[i]]]
  #If a measurement exists for the animal and is not an empty dictionary
  #lookup the result string by UAR and measurement ID and insert into spreadsheet
  for j in range (0, len(meas_by_day)):
    if results[animal_uars[i]][meas_by_day[j]] != None and results[animal_uars[i]][meas_by_day[j]] != {}:
      ws.cell(row=i+2, column=j+3).value = results[animal_uars[i]][meas_by_day[j]]['RESULT_STRING']

#Insert row/column headers and format the group summary table layout
for x in range (0, len(grp_meas_by_day)):
  ws2.merge_cells(start_row=(((x+1)*7+x)-6),start_column=2,end_row=(((x+1)*7+x)-6),end_column=len(group_grp_ids)+1)
  ws2.cell(row=(((x+1)*7+x)-6), column=2).value = grp_meas_by_day[x]
  for j in range (0, len(group_grp_ids)):
    ws2.cell(row=(((x+1)*7+x)-5), column=j+2).value = group_grp_key[group_grp_ids[j]]
  for i in range (0, len(grp_summary_ids)):
    ws2.cell(row=(((x+1)*7+x)-4)+i, column=1).value = grp_summary_key[grp_summary_ids[i]]

#Put data into the group layout
for x in range (0, len(grp_meas_by_day)):
  for y in range (0, len(group_grp_ids)):
    for z in range (0, len(grp_summary_ids)):
      if grp_results[grp_meas_by_day[x]][group_grp_ids[y]][grp_summary_ids[z]]['RESULT_STRING'] != {}:
        ws2.cell(row=(((x+1)*7+x+z)-4), column=y+2).value = grp_results[grp_meas_by_day[x]][group_grp_ids[y]][grp_summary_ids[z]]['RESULT_STRING']


save_name = (sys.argv[1].split("."))[0]
      
wb.save('%s.xlsx' % save_name)